"""Re-grade entire class with current code, overwrite Excel."""
import sys, os, yaml, glob
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from src.extractor import extract, extract_from_student_folder
from src.preprocessor import process
from src.utils import load_rubric
from src.grader import grade
import openpyxl
import argparse

parser = argparse.ArgumentParser()
parser.add_argument("class_dir")
parser.add_argument("--strategy", default="free_vision")
args = parser.parse_args()

class_dir = args.class_dir
class_name = os.path.basename(class_dir)

with open("config.yaml", "r", encoding="utf-8") as f:
    cfg = yaml.safe_load(f)
rubric = load_rubric("data/rubric.json")
config = {"vision_strategy": args.strategy, "grading": cfg.get("grading",{}),
          "image": cfg.get("image",{}), "model_router": cfg.get("model_router",{})}

# Collect all students
students = []
for entry in sorted(os.listdir(class_dir)):
    fp = os.path.join(class_dir, entry)
    if not os.path.isdir(fp): continue

    # Find docx
    docx_files = []
    for f in os.listdir(fp):
        if f.startswith('~$'): continue
        fl = f.lower()
        if fl.endswith('.docx') or (fl.endswith('docx') and not fl.endswith('.docx')):
            docx_files.append(os.path.join(fp, f))
    if not docx_files: continue

    # Prefer student submission over template
    student_docx = [d for d in docx_files if not any(kw in os.path.basename(d) for kw in ['试卷','期末','《'])]
    if not student_docx: student_docx = docx_files
    docx = max(student_docx, key=lambda x: os.path.getsize(x))

    # Extract student info
    sid = entry.split('-')[0] if '-' in entry else ''
    if not sid.isdigit():
        import re
        m = re.match(r'(\d+)', entry)
        sid = m.group(1) if m else entry[:12]
    name = entry[len(sid):].lstrip('-') if entry.startswith(sid) else entry

    # Collect supplementary files
    images = []
    excels = []
    for f in os.listdir(fp):
        ff = os.path.join(fp, f)
        if f.lower().endswith(('.png','.jpg','.jpeg')):
            images.append(ff)
        elif f.lower().endswith(('.xlsx','.xls')):
            excels.append(ff)

    students.append({"docx": docx, "sid": sid, "name": name, "images": images, "excels": excels})

print(f"Class: {class_name}  Students: {len(students)}  Strategy: {args.strategy}")

# Grade each student
results = []
for i, s in enumerate(students):
    try:
        paper = extract_from_student_folder(s["docx"], "output/test",
                                            {"images": s["images"], "excel": s["excels"]})
        clean = process(paper, rubric, config)
    except Exception as e:
        print(f"  [{i+1}] {s['name']}: EXTRACT ERROR {e}")
        continue

    all_criteria = []
    q_totals = []
    total = 0
    for q in rubric["questions"]:
        qk = f"q{q['id']}"
        qd = clean.get(qk, {})
        try:
            r = grade(qd, q, config)
            sc = r.get("总分", 0)
            total += sc
            q_totals.append(sc)
            for c in q["criteria"]:
                key = "得分_" + c["id"] + "_" + c["name"]
                all_criteria.append(r.get(key, 0))
        except Exception as e:
            q_totals.append(0)
            for c in q["criteria"]:
                all_criteria.append(0)

    print(f"  [{i+1}] {s['name']:<8} {total:>3}  Q1={q_totals[0]} Q2={q_totals[1]} Q3={q_totals[2]} Q4={q_totals[3]} Q5={q_totals[4]}")
    results.append({"name": s["name"], "sid": s["sid"], "total": total,
                    "criteria": all_criteria, "q_totals": q_totals})

# Write Excel
xlsx_path = f"output/{class_name}/评分汇总_{class_name}.xlsx"
os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "得分表"

# Header rows
title = f"人工智能基础实践 — {class_name} 得分表"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=24)
ws.cell(row=1, column=1, value=title)

headers = ["序号","学号","姓名",
           "Q1 文本生成(15分)","","",
           "Q2 图像设计(20分)","","","",
           "Q3 视频创作(20分)","","","","",
           "Q4 数据处理(20分)","","","",
           "Q5 智能体搭建(25分)","","","",
           "总分"]
ws.append(headers)

sub_headers = ["","","",
               "主题契合度","文案内容适配","提示词与提交完整性",
               "主题契合度","工具使用的合理性","设计图像的创意和完整性","提示词与提交完整性",
               "主题契合度与文化准确性","图片生成质量","人物服饰与背景","配音与整体效果","提交完整性",
               "缺失值处理","重复项处理","格式统一与排序","提交完整性",
               "人设与回复逻辑","知识库配置","智能体功能完善性","提交完整性"]
ws.append(sub_headers)

max_scores = ["","","",5,5,5, 5,5,5,5, 4,4,4,4,4, 5,5,5,5, 7,7,6,5]
ws.append(max_scores)

# Data rows
for i, r in enumerate(results):
    row = [i+1, r["sid"], r["name"]] + r["criteria"] + [r["total"]]
    ws.append(row)

# Average row
avg_row = len(results) + 5
ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=3)
ws.cell(row=avg_row, column=1, value="平均")
for col in range(4, 25):
    total_val = 0
    count = 0
    for r in results:
        idx = col - 4
        if idx < len(r["criteria"]):
            total_val += r["criteria"][idx]
            count += 1
    ws.cell(row=avg_row, column=col, value=round(total_val/count, 1) if count else "")

wb.save(xlsx_path)

# Stats
totals = [r["total"] for r in results]
avg = sum(totals)//len(totals) if totals else 0
print(f"\nDone: {xlsx_path}")
print(f"Avg: {avg}  Min: {min(totals)}  Max: {max(totals)}")
