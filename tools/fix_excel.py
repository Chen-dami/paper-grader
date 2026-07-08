"""Fix 电商2501 grading Excel: update mis-scored + add 杨宝妹"""
import openpyxl
from copy import copy

xlsx_path = r'C:\Users\Dami\Desktop\ai-grader\output\电商2501\评分汇总_电商2501.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# ============================================================
# 1. 新评分数据 (用当前代码重评)
# ============================================================
# Format: sid -> {row_data: [q1_c1, q1_c2, q1_c3, q2_c1..q2_c4, q3_c1..q3_c5, q4_c1..q4_c4, q5_c1..q5_c4]}
updates = {
    '255307010108': [5,5,5, 5,5,5,5, 4,4,4,4,2, 5,5,5,5, 7,7,6,5],   # 符大靖 Q5:0→25
    '255307010110': [5,5,5, 5,5,5,5, 4,4,4,4,4, 4,4,4,5, 7,7,6,5],   # 吴俊杰 Q5:0→25
    '255307010103': [5,5,5, 0,5,0,5, 4,4,4,4,2, 0,0,0,0, 7,0,6,5],   # 吉才创 Q2:0→10 Q5:7→18
    '255307010106': [5,5,5, 5,5,5,5, 4,4,4,0,2, 1,1,1,1, 7,0,0,0],   # 黄良圳 Q2:0→20
    '255307010109': [5,5,5, 0,0,0,0, 4,4,4,4,4, 5,5,5,5, 7,7,6,5],   # 陈振哲 Q5:14→25
    '255307010126': [5,5,5, 5,5,5,5, 4,4,4,4,4, 5,5,5,5, 0,0,0,0],   # 杨宝妹 新增
}

# ============================================================
# 2. 更新已有学生的分数
# ============================================================
updated_count = 0
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    sid_cell = row[1]  # 学号列
    if sid_cell.value is None: continue
    sid = str(sid_cell.value)
    if sid in updates and sid != '255307010126':  # 126 is new
        new_scores = updates[sid]
        # Write sub-scores: cols 3-22 (0-indexed: 3 to 22)
        for i, val in enumerate(new_scores):
            row[3 + i].value = val
        # Recalculate total: cols 3-22
        total = sum(new_scores)
        row[23].value = total  # Total column (0-indexed: 23)
        updated_count += 1
        name = str(row[2].value) if row[2].value else '?'
        print(f'Updated: {sid} {name} -> total={total}')

# ============================================================
# 3. 添加杨宝妹为新行
# ============================================================
new_scores = updates['255307010126']
total_ym = sum(new_scores)

# Find the insert position (before the average row)
insert_row = None
for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), start=4):
    if row[1].value and str(row[1].value).strip().startswith('平均'):
        insert_row = i
        break
if insert_row is None:
    insert_row = ws.max_row  # Before average

# Get the new 序号
prev_num = 0
for row in ws.iter_rows(min_row=4, max_row=insert_row-1):
    if row[0].value and str(row[0].value).isdigit():
        prev_num = max(prev_num, int(row[0].value))
new_num = prev_num + 1

ws.insert_rows(insert_row)
new_row = insert_row

# Copy style from previous row
src_row = insert_row - 1

# Write data
ws.cell(row=new_row, column=1, value=new_num)  # 序号
ws.cell(row=new_row, column=2, value='255307010126')  # 学号
ws.cell(row=new_row, column=3, value='杨宝妹')  # 姓名
for i, val in enumerate(new_scores):
    ws.cell(row=new_row, column=4 + i, value=val)
ws.cell(row=new_row, column=24, value=total_ym)  # 总分

print(f'Added: 255307010126 杨宝妹 -> total={total_ym} at row {new_row}')

# ============================================================
# 4. 更新平均行
# ============================================================
avg_row = None
for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row+1), start=4):
    if row[1].value and '平均' in str(row[1].value):
        avg_row = i
        break

if avg_row:
    for col in range(4, 24):  # cols 4-23 (1-indexed)
        total_val = 0
        count = 0
        for row in ws.iter_rows(min_row=4, max_row=avg_row-1, min_col=col, max_col=col):
            v = row[0].value
            if isinstance(v, (int, float)):
                total_val += v
                count += 1
        if count > 0:
            ws.cell(row=avg_row, column=col, value=round(total_val / count, 1))
    print(f'Updated averages at row {avg_row}')

# ============================================================
# Save
# ============================================================
backup = xlsx_path.replace('.xlsx', '_backup.xlsx')
import shutil
shutil.copy2(xlsx_path, backup)
wb.save(xlsx_path)
print(f'\nSaved: {xlsx_path}')
print(f'Backup: {backup}')
print(f'Updated {updated_count} students + added 杨宝妹')
