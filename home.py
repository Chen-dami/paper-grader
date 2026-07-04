"""
主页 —— 系统仪表盘
"""
import streamlit as st
import os, sys, json, time, glob

sys.path.insert(0, os.path.dirname(__file__))
from src.ui_style import inject; inject()

if not st.session_state.get("authenticated", False):
	st.warning("请先登录")
	st.stop()

teacher = st.session_state.teacher_name

# ============================================================
#  收集系统状态
# ============================================================

# 评分标准
rubric_info = None
rubric_path = st.session_state.get("rubric_path", "data/rubric.json")
if os.path.exists(rubric_path):
	try:
		with open(rubric_path, "r", encoding="utf-8") as f:
			r = json.load(f)
		rubric_info = {
			"name": r.get("exam", {}).get("name", "未知"),
			"total": r.get("exam", {}).get("total_score", 0),
			"questions": len(r.get("questions", [])),
		}
	except:
		pass

# 试卷统计
total_papers = 0
class_count = 0
if os.path.exists("data/papers"):
	for d in os.listdir("data/papers"):
		dpath = os.path.join("data/papers", d)
		if os.path.isdir(dpath):
			docx_count = len(glob.glob(os.path.join(dpath, "**", "*.docx"), recursive=True))
			if docx_count > 0:
				class_count += 1
				total_papers += docx_count

# 阅卷记录
output_count = 0
if os.path.exists("output"):
	for d in os.listdir("output"):
		if os.path.isdir(os.path.join("output", d)):
			summary = os.path.join("output", d, f"评分汇总_{d}.xlsx")
			if os.path.exists(summary):
				output_count += 1

# ============================================================
#  页面
# ============================================================
st.title("阅卷系统")
st.caption(f"{teacher}，{time.strftime('%Y年%m月%d日 %H:%M')} | AI 驱动的智能阅卷系统")

st.divider()

# ===== 状态卡片 =====
st.subheader("系统概览")
mc = st.columns(4)
with mc[0]:
	st.metric("当前考试", rubric_info["name"] if rubric_info else "未设置",
	          delta=f"{rubric_info['total']}分 · {rubric_info['questions']}题" if rubric_info else None)
with mc[1]:
	st.metric("待阅班级", class_count,
	          delta=f"{total_papers} 份试卷" if total_papers > 0 else None)
with mc[2]:
	st.metric("已完成阅卷", output_count,
	          delta=f"共 {output_count} 个班级" if output_count > 0 else None)
with mc[3]:
	mode = st.session_state.get("grading_mode", "relaxed")
	mode_label = {"relaxed": "宽松", "normal": "标准", "strict": "严格"}.get(mode, mode)
	st.metric("评分模式", mode_label)

st.divider()

# ===== 快捷入口 =====
st.subheader("快捷入口")
col1, col2, col3, col4 = st.columns(4)

with col1:
	with st.container(border=True):
		st.markdown("### 开始阅卷")
		st.caption("上传评分标准 → 选择班级 → 一键批改")
		if st.button("进入阅卷 →", type="primary", use_container_width=True, key="goto_grade"):
			st.switch_page("pages/1_grading.py")

with col2:
	with st.container(border=True):
		st.markdown("### AI 助手")
		st.caption("对话式修改评分标准和配置")
		if st.button("打开助手 →", use_container_width=True, key="goto_ai"):
			st.switch_page("pages/4_ai_assistant.py")

with col3:
	with st.container(border=True):
		st.markdown("### 查看结果")
		st.caption("历史成绩、统计图表、下载报告")
		if st.button("查看历史 →", use_container_width=True, key="goto_results"):
			st.switch_page("pages/3_results.py")

with col4:
	with st.container(border=True):
		st.markdown("### 学生查询")
		st.caption("按学号快速查找学生成绩")
		if st.button("查询成绩 →", use_container_width=True, key="goto_student"):
			st.switch_page("pages/6_student_portal.py")

st.divider()

# ===== 最近班级状态 =====
st.subheader("班级状态")

if os.path.exists("output") and output_count > 0:
	class_dirs = sorted(
		[d for d in os.listdir("output") if os.path.isdir(os.path.join("output", d))],
		reverse=True
	)
	status_data = []
	for d in class_dirs[:10]:
		dpath = os.path.join("output", d)
		summary = os.path.join(dpath, f"评分汇总_{d}.xlsx")
		check = os.path.join(dpath, "查重报告.xlsx")

		status = "✅ 已完成"
		has_check = "✅" if os.path.exists(check) else "—"

		student_count = "?"
		if os.path.exists(summary):
			try:
				from openpyxl import load_workbook
				wb = load_workbook(summary, data_only=True)
				ws = wb[wb.sheetnames[0]]
				count = 0
				for row in ws.iter_rows(min_row=5, values_only=True):
					if row and row[0] and str(row[0]).strip() not in ("", "平均", "平均分", "None", "nan"):
						count += 1
				student_count = str(count)
			except:
				pass

		status_data.append({
			"班级": d,
			"状态": status,
			"学生数": student_count,
			"查重": has_check,
		})

	if status_data:
		import pandas as pd
		st.dataframe(pd.DataFrame(status_data), use_container_width=True, hide_index=True,
		             height=min(400, 35 * len(status_data) + 38))
else:
	st.info("还没有阅卷记录。" if total_papers == 0 else f"有 {total_papers} 份试卷待阅卷，去「阅卷」页开始吧！")
