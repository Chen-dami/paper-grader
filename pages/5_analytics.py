"""
数据分析 -- ECharts 可视化 + 统计指标
"""
import streamlit as st
import os, sys, json
import pandas as pd
import numpy as np
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from src.ui_style import inject; inject()

if not st.session_state.get("authenticated", False):
    st.warning("请先登录")
    st.stop()

st.title("数据分析")

# ============================================================
#  ECharts HTML 模板
# ============================================================
ECHARTS_CDN = "https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"

def echarts_html(option_json: str, height: int = 350) -> str:
    return f"""
    <div id="chart" style="width:100%;height:{height}px;"></div>
    <script src="{ECHARTS_CDN}"></script>
    <script>
    var chart = echarts.init(document.getElementById('chart'));
    chart.setOption({option_json});
    window.addEventListener('resize', function(){{chart.resize()}});
    </script>
    """

def render_echarts(option: dict, height: int = 350):
    import json as _json
    opt_str = _json.dumps(option, ensure_ascii=False, default=str)
    st.components.v1.html(echarts_html(opt_str, height), height=height + 20)

# ============================================================
#  数据加载
# ============================================================
def scan_classes():
    classes = []
    if not os.path.exists("output"):
        return classes
    for d in os.listdir("output"):
        sp = os.path.join("output", d, f"评分汇总_{d}.xlsx")
        if os.path.isdir(os.path.join("output", d)) and os.path.exists(sp):
            classes.append({"name": d, "summary": sp})
    return sorted(classes, key=lambda x: x["name"])

@st.cache_data(ttl=30)
def load_all_data():
    all_data = {}
    for cls in scan_classes():
        try:
            wb = load_workbook(cls["summary"], data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if len(rows) < 5:
                continue
            row2, row3, row4 = rows[1], rows[2], rows[3]
            max_len = max(len(row2), len(row3), len(row4))
            questions = []
            cur_q, cur_items = None, []
            for i in range(3, max_len):
                r2 = str(row2[i]).strip() if i < len(row2) and row2[i] else ""
                r3 = str(row3[i]).strip() if i < len(row3) and row3[i] else ""
                r4 = row4[i] if i < len(row4) and row4[i] else 0
                if "总分" in (r2 + r3):
                    break
                if r2 and r2 not in ("None", ""):
                    if cur_q:
                        questions.append({"name": cur_q, "items": cur_items})
                    cur_q = r2.replace("\n", " ")
                    cur_items = []
                if r3 and r3 not in ("None", ""):
                    cmax = int(r4) if isinstance(r4, (int, float)) and not isinstance(r4, bool) else 0
                    cur_items.append({"name": r3.replace("\n", " "), "max": cmax})
            if cur_q:
                questions.append({"name": cur_q, "items": cur_items})

            col_map = ["序号", "学号", "姓名"]
            for q in questions:
                for item in q["items"]:
                    col_map.append(f"{q['name']}_{item['name']}")
            col_map.append("总分")

            data_rows = []
            for row in rows[4:]:
                vals = list(row) if row else []
                if not vals or all(v is None for v in vals):
                    continue
                first = str(vals[0]).strip() if vals[0] else ""
                if first in ("", "平均", "平均分", "None", "nan"):
                    continue
                while len(vals) < len(col_map):
                    vals.append(None)
                vals = vals[:len(col_map)]
                data_rows.append(vals)

            if not data_rows:
                continue
            df = pd.DataFrame(data_rows, columns=col_map)
            for c in df.columns[3:]:
                df[c] = pd.to_numeric(df[c], errors="coerce")
            all_data[cls["name"]] = {"df": df, "questions": questions}
        except Exception:
            continue
    return all_data

all_data = load_all_data()
if not all_data:
    st.info("还没有阅卷记录")
    st.stop()

# ============================================================
#  统计算法
# ============================================================
def calc_difficulty(scores, max_score):
    valid = scores.dropna()
    if len(valid) == 0 or max_score == 0:
        return 0, "N/A"
    p = valid.mean() / max_score
    if p >= 0.85: level = "简单"
    elif p >= 0.6: level = "适中"
    elif p >= 0.4: level = "偏难"
    else: level = "困难"
    return round(p, 3), level

def calc_discrimination(scores, total_scores, max_score):
    valid = pd.DataFrame({"score": scores, "total": total_scores}).dropna()
    if len(valid) < 4 or max_score == 0:
        return 0, "N/A"
    n = len(valid)
    top_n = max(1, n // 3)
    sdf = valid.sort_values("total", ascending=False)
    high = sdf.head(top_n)["score"].mean()
    low = sdf.tail(top_n)["score"].mean()
    d = (high - low) / max_score
    if d >= 0.4: level = "优秀"
    elif d >= 0.3: level = "良好"
    elif d >= 0.2: level = "尚可"
    else: level = "需改进"
    return round(d, 3), level

def calc_skewness_kurtosis(scores):
    valid = scores.dropna()
    if len(valid) < 3:
        return 0, 0
    n = len(valid)
    mean = valid.mean()
    std = valid.std(ddof=0)
    if std == 0:
        return 0, 0
    skew = (n / ((n - 1) * (n - 2))) * sum(((valid - mean) / std) ** 3)
    kurt = ((n * (n + 1)) / ((n - 1) * (n - 2) * (n - 3))) * sum(((valid - mean) / std) ** 4)
    kurt -= (3 * (n - 1) ** 2) / ((n - 2) * (n - 3))
    return round(skew, 3), round(kurt, 3)

# ============================================================
#  班级选择
# ============================================================
class_names = list(all_data.keys())
selected = st.selectbox("选择班级", class_names, key="anal_class")
data = all_data[selected]
df = data["df"]
questions = data["questions"]

q_totals = {}
for q in questions:
    cols = [f"{q['name']}_{item['name']}" for item in q["items"]]
    valid_cols = [c for c in cols if c in df.columns]
    if valid_cols:
        df[f"__{q['name']}__"] = df[valid_cols].sum(axis=1)
        q_totals[q["name"]] = sum(item["max"] for item in q["items"])

total_col = "总分"
has_total = total_col in df.columns

# ============================================================
#  概览指标
# ============================================================
st.subheader("概览")
mc = st.columns(6)
with mc[0]:
    st.metric("学生数", len(df))
with mc[1]:
    st.metric("平均分", f"{df[total_col].mean():.1f}" if has_total else "N/A")
with mc[2]:
    st.metric("中位数", f"{df[total_col].median():.1f}" if has_total else "N/A")
with mc[3]:
    st.metric("标准差", f"{df[total_col].std():.1f}" if has_total else "N/A")
with mc[4]:
    st.metric("及格率", f"{(df[total_col]>=60).sum()/len(df)*100:.0f}%" if has_total else "N/A")
with mc[5]:
    if has_total:
        skew, kurt = calc_skewness_kurtosis(df[total_col])
        st.metric("偏度/峰度", f"{skew}/{kurt}")

# ============================================================
#  图表行1：成绩分布 + 各题得分率
# ============================================================
st.divider()
row1_l, row1_r = st.columns(2)

with row1_l:
    st.subheader("成绩分布")
    if has_total:
        scores = df[total_col].dropna()
        bins_val = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
        hist, _ = np.histogram(scores, bins=bins_val)
        cats = ["0-10","10-20","20-30","30-40","40-50","50-60","60-70","70-80","80-90","90-100"]
        opt = {
            "tooltip": {"trigger": "axis"},
            "xAxis": {"data": cats, "axisLabel": {"fontSize": 10}},
            "yAxis": {"type": "value", "name": "人数"},
            "series": [{"type": "bar", "data": [int(x) for x in hist],
                        "itemStyle": {"color": "#4A7C59"},
                        "markLine": {"data": [{"yAxis": float(scores.mean()), "name": f"均值{scores.mean():.1f}",
                                              "lineStyle": {"color": "#C0392B", "type": "dashed"}}]}}],
            "grid": {"left": 40, "right": 20, "top": 20, "bottom": 30},
        }
        render_echarts(opt)

with row1_r:
    st.subheader("各题得分率")
    if q_totals:
        rates = []
        for qname, qmax in q_totals.items():
            col = f"__{qname}__"
            if col in df.columns and qmax > 0:
                avg = df[col].mean()
                rates.append((qname[:20], round(avg / qmax * 100, 1)))
        if rates:
            rates.sort(key=lambda x: x[1])
            opt = {
                "tooltip": {"trigger": "axis"},
                "xAxis": {"type": "value", "name": "得分率(%)", "max": 100},
                "yAxis": {"type": "category", "data": [r[0] for r in rates], "inverse": True,
                          "axisLabel": {"fontSize": 10}},
                "series": [{"type": "bar", "data": [r[1] for r in rates],
                            "itemStyle": {"color": "#4A7C59"},
                            "label": {"show": True, "position": "right",
                                      "formatter": "{c}%", "fontSize": 10}}],
                "grid": {"left": 5, "right": 45, "top": 10, "bottom": 20},
            }
            render_echarts(opt)

# ============================================================
#  图表行2：等级分布 + 成绩排名曲线
# ============================================================
row2_l, row2_r = st.columns(2)

with row2_l:
    st.subheader("等级分布")
    if has_total:
        levels = {"A (>=90)": 0, "B (80-89)": 0, "C (70-79)": 0, "D (60-69)": 0, "F (<60)": 0}
        for s in df[total_col].dropna():
            if s >= 90: levels["A (>=90)"] += 1
            elif s >= 80: levels["B (80-89)"] += 1
            elif s >= 70: levels["C (70-79)"] += 1
            elif s >= 60: levels["D (60-69)"] += 1
            else: levels["F (<60)"] += 1
        pie_data = [{"name": k, "value": v} for k, v in levels.items() if v > 0]
        opt = {
            "tooltip": {"trigger": "item", "formatter": "{b}: {c}人 ({d}%)"},
            "series": [{"type": "pie", "radius": ["40%", "70%"],
                        "data": pie_data,
                        "label": {"formatter": "{b}\n{d}%"}},
                       ],
        }
        render_echarts(opt)

with row2_r:
    st.subheader("成绩排名曲线")
    if has_total:
        sorted_scores = sorted(df[total_col].dropna(), reverse=True)
        opt = {
            "tooltip": {"trigger": "axis"},
            "xAxis": {"type": "category", "name": "排名"},
            "yAxis": {"type": "value", "name": "分数"},
            "series": [{"type": "line", "data": sorted_scores,
                        "smooth": True, "lineStyle": {"color": "#4A7C59", "width": 2},
                        "symbol": "none"}],
            "grid": {"left": 50, "right": 15, "top": 10, "bottom": 30},
        }
        render_echarts(opt)

# ============================================================
#  题目质量分析
# ============================================================
st.divider()
st.subheader("题目质量分析")

if q_totals and has_total:
    quality_rows = []
    for qname, qmax in q_totals.items():
        col = f"__{qname}__"
        if col in df.columns:
            sc = df[col].dropna()
            diff, dl = calc_difficulty(sc, qmax)
            disc, dcl = calc_discrimination(sc, df[total_col], qmax)
            quality_rows.append({
                "题目": qname[:30], "满分": qmax,
                "平均分": round(sc.mean(), 1),
                "难度系数P": diff, "难度评价": dl,
                "区分度D": disc, "区分度评价": dcl,
                "标准差": round(sc.std(), 2),
            })
    if quality_rows:
        qdf = pd.DataFrame(quality_rows)
        st.dataframe(qdf, use_container_width=True, hide_index=True)

        # 难度-区分度散点图
        scatter_data = []
        for _, r in qdf.iterrows():
            scatter_data.append({
                "value": [r["难度系数P"], r["区分度D"]],
                "name": r["题目"][:12],
            })
        opt = {
            "tooltip": {"formatter": "function(p){return p.name+'<br/>难度:'+p.value[0].toFixed(3)+' 区分度:'+p.value[1].toFixed(3)}"},
            "xAxis": {"type": "value", "name": "难度系数 P", "min": 0, "max": 1.05},
            "yAxis": {"type": "value", "name": "区分度 D", "min": -0.1, "max": 1.05},
            "series": [{"type": "scatter", "data": scatter_data,
                        "symbolSize": 16,
                        "itemStyle": {"color": "#4A7C59"},
                        "label": {"show": True, "position": "top", "fontSize": 10,
                                  "formatter": "{b}"},
                        "markLine": {"data": [
                            {"yAxis": 0.3, "name": "良好", "lineStyle": {"color": "#27AE60", "type": "dashed"}},
                            {"yAxis": 0.2, "name": "尚可", "lineStyle": {"color": "#E67E22", "type": "dashed"}},
                        ]}}],
            "grid": {"left": 55, "right": 20, "top": 20, "bottom": 40},
        }
        render_echarts(opt, height=420)

# ============================================================
#  班级对比
# ============================================================
if len(all_data) > 1:
    st.divider()
    st.subheader("班级对比")

    compare_rows = []
    for cname, cdata in all_data.items():
        cdf = cdata["df"]
        if total_col in cdf.columns:
            sc = cdf[total_col].dropna()
            if len(sc) > 0:
                compare_rows.append({
                    "班级": cname, "人数": len(sc),
                    "平均分": round(sc.mean(), 1),
                    "中位数": round(sc.median(), 1),
                    "标准差": round(sc.std(), 2),
                    "及格率(%)": round((sc >= 60).sum() / len(sc) * 100, 1),
                    "优秀率(%)": round((sc >= 90).sum() / len(sc) * 100, 1),
                })

    if compare_rows:
        cpdf = pd.DataFrame(compare_rows)
        st.dataframe(cpdf, use_container_width=True, hide_index=True)

        opt = {
            "tooltip": {"trigger": "axis"},
            "xAxis": {"type": "category", "data": cpdf["班级"].tolist()},
            "yAxis": {"type": "value", "name": "平均分", "max": 100},
            "series": [{"type": "bar", "data": cpdf["平均分"].tolist(),
                        "itemStyle": {"color": "#4A7C59"},
                        "label": {"show": True, "position": "top", "fontSize": 11,
                                  "formatter": "{c}"}}],
            "grid": {"left": 40, "right": 20, "top": 10, "bottom": 50},
        }
        render_echarts(opt, height=300)

# ============================================================
#  成绩明细
# ============================================================
st.divider()
st.subheader("成绩明细")
show_cols = ["学号", "姓名"]
for q in questions:
    col = f"__{q['name']}__"
    if col in df.columns:
        show_cols.append(col)
show_cols.append("总分")
show_cols = [c for c in show_cols if c in df.columns]
rename = {}
for q in questions:
    col = f"__{q['name']}__"
    if col in df.columns:
        rename[col] = q["name"][:18]
display_df = df[show_cols].rename(columns=rename).sort_values("总分", ascending=False)
st.dataframe(display_df, use_container_width=True, hide_index=True,
             height=min(500, 35 * len(df) + 38))

csv = display_df.to_csv(index=False).encode("utf-8-sig")
st.download_button("导出 CSV", csv, f"{selected}_成绩.csv", mime="text/csv")
