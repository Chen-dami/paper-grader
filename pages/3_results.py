"""
历史结果页
"""
import streamlit as st
import os, sys, glob, io, zipfile
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from src.ui_style import inject; inject()

if not st.session_state.get("authenticated", False):
	st.warning("请先在首页登录")
	st.stop()

st.title("历史结果")

if not os.path.exists("output"):
	st.info("还没有阅卷记录")
	st.stop()

class_dirs = []
for d in os.listdir("output"):
	path = os.path.join("output", d)
	if os.path.isdir(path) and os.path.exists(os.path.join(path, f"评分汇总_{d}.xlsx")):
		class_dirs.append(d)

if not class_dirs:
	st.info("output/ 下没有找到评分汇总")
	st.stop()

selected = st.selectbox("选择班级", class_dirs, index=len(class_dirs) - 1)
out_dir = os.path.join("output", selected)
summary_path = os.path.join(out_dir, f"评分汇总_{selected}.xlsx")
check_path = os.path.join(out_dir, "查重报告.xlsx")
personal_dir = os.path.join(out_dir, "个人成绩")

@st.cache_data(ttl=10)
def load_summary(path):
	"""解析班级汇总Excel，正确处理三行合并表头及合并单元格"""
	wb = load_workbook(path)
	ws = wb[wb.sheetnames[0]]

	all_rows = list(ws.iter_rows(min_row=1, values_only=True))
	if len(all_rows) < 5:
		return pd.DataFrame()

	# 行2-4是三行表头：大题名 / 得分项名 / 满分值（0-index: 1,2,3）
	row2 = all_rows[1] if len(all_rows) > 1 else []
	row3 = all_rows[2] if len(all_rows) > 2 else []
	row4 = all_rows[3] if len(all_rows) > 3 else []

	# 构建唯一列名
	headers = []
	max_len = max(len(row2), len(row3), len(row4))
	for i in range(max_len):
		r2 = str(row2[i]).strip() if i < len(row2) and row2[i] is not None else ""
		r3 = str(row3[i]).strip() if i < len(row3) and row3[i] is not None else ""
		r4 = row4[i] if i < len(row4) and row4[i] is not None else ""

		if r3 in ("", "None") and r2 in ("", "None"):
			# 固定列：序号 / 学号 / 姓名
			if i == 0:
				headers.append("序号")
			elif i == 1:
				headers.append("学号")
			elif i == 2:
				headers.append("姓名")
			else:
				# 可能是总分列
				headers.append(f"col_{i}")
		elif r3 in ("", "None"):
			if "总分" in r2 or "总" in r2:
				headers.append("总分")
			else:
				headers.append(r2 if r2 else f"col_{i}")
		else:
			# 得分项：大题_得分项(满分)
			q_part = r2.replace("\n", " ").strip() if r2 else ""
			c_part = r3.replace("\n", " ").strip() if r3 else ""
			max_val = int(r4) if isinstance(r4, (int, float)) and not isinstance(r4, bool) else ""
			max_part = f"({max_val})" if max_val != "" else ""
			if q_part and c_part:
				headers.append(f"{q_part}_{c_part}{max_part}")
			elif c_part:
				headers.append(f"{c_part}{max_part}")
			else:
				headers.append(f"col_{i}")

	# 去重：给重复列名加后缀
	seen = {}
	unique_headers = []
	for h in headers:
		if h in seen:
			seen[h] += 1
			unique_headers.append(f"{h}_{seen[h]}")
		else:
			seen[h] = 0
			unique_headers.append(h)

	# 数据从第5行开始
	data = []
	for row in all_rows[4:]:
		vals = list(row) if row else []
		if vals and any(v is not None for v in vals):
			if len(vals) < len(unique_headers):
				vals = list(vals) + [None] * (len(unique_headers) - len(vals))
			elif len(vals) > len(unique_headers):
				vals = vals[:len(unique_headers)]
			data.append(vals)

	if not data:
		return pd.DataFrame()

	df = pd.DataFrame(data, columns=unique_headers)

	# 移除平均行
	if len(df) > 0:
		first_col = df.iloc[:, 0].astype(str).str.strip()
		df = df[~first_col.isin(["", "平均", "平均分", "None", "nan"])].copy()

	return df


df = load_summary(summary_path)
if df is None or df.empty:
	st.warning("无法加载成绩数据")
	st.stop()

# 安全数值转换：只对数据列操作，跳过非Series和文本列
skip_patterns = ["序号", "学号", "姓名", "等级"]
for c in list(df.columns):
	if any(p in str(c) for p in skip_patterns):
		continue
	col_data = df[c]
	if hasattr(col_data, 'ndim') and col_data.ndim != 1:
		continue
	try:
		numeric_col = pd.to_numeric(col_data, errors="coerce")
		if numeric_col.notna().sum() >= len(numeric_col) * 0.5:
			df[c] = numeric_col
	except (TypeError, ValueError):
		pass

total_col = "总分"
if total_col not in df.columns:
	# 找到总分列（包含"总分"的列名）
	tc = [c for c in df.columns if "总分" in str(c)]
	total_col = tc[0] if tc else (list(df.columns)[-1] if len(df.columns) > 0 else None)

st.subheader(f"{selected} -- 成绩概览")
mc = st.columns(5)
with mc[0]: st.metric("学生数", len(df))
with mc[1]:
	if total_col and total_col in df.columns:
		st.metric("平均分", f"{df[total_col].mean():.1f}")
with mc[2]:
	if total_col and total_col in df.columns:
		st.metric("最高分", f"{df[total_col].max():.0f}")
with mc[3]:
	if total_col and total_col in df.columns:
		st.metric("最低分", f"{df[total_col].min():.0f}")
with mc[4]:
	if total_col and total_col in df.columns:
		pr = (df[total_col] >= 60).sum() / len(df) * 100
		st.metric("及格率", f"{pr:.0f}%")

st.divider()
cc1, cc2 = st.columns(2)
with cc1:
	st.subheader("分数段分布")
	if total_col and total_col in df.columns:
		scores = df[total_col].dropna()
		bands = {"90-100": 0, "80-89": 0, "70-79": 0, "60-69": 0, "<60": 0}
		for s in scores:
			if s >= 90: bands["90-100"] += 1
			elif s >= 80: bands["80-89"] += 1
			elif s >= 70: bands["70-79"] += 1
			elif s >= 60: bands["60-69"] += 1
			else: bands["<60"] += 1
		st.bar_chart(pd.DataFrame({"分数段": list(bands.keys()), "人数": list(bands.values())}).set_index("分数段"),
		             use_container_width=True)
with cc2:
	st.subheader("各题平均得分率")
	# 找到所有含"(数字)"的得分列
	q_cols = [c for c in df.columns if "总分" not in str(c)
	          and any(p not in str(c) for p in skip_patterns)
	          and "col_" not in str(c)]
	if q_cols:
		import re
		q_rates = {}
		for qc in q_cols:
			if qc in df.columns and pd.api.types.is_numeric_dtype(df[qc]):
				mm = re.search(r'\((\d+)\)', str(qc))
				qmax = int(mm.group(1)) if mm else df[qc].max()
				if qmax > 0:
					q_rates[qc] = df[qc].mean() / qmax * 100
		if q_rates:
			st.bar_chart(pd.DataFrame({"题目": list(q_rates.keys()), "得分率(%)": list(q_rates.values())}).set_index("题目"),
			             use_container_width=True)

st.divider()
st.subheader("学生成绩表")
st.dataframe(df, use_container_width=True, hide_index=True, height=min(500, 35 * len(df) + 38))

st.divider()
st.subheader("个人明细")
if os.path.exists(personal_dir):
	pf = sorted([f for f in os.listdir(personal_dir) if f.endswith(".xlsx")])
	if pf:
		sf = st.selectbox("选择学生", pf)
		if sf:
			wb = load_workbook(os.path.join(personal_dir, sf))
			ws = wb.active
			for row in ws.iter_rows(values_only=True):
				vals = [str(v) if v is not None else "" for v in row]
				text = " | ".join(v for v in vals if v)
				if text.strip():
					st.text(text)

st.divider()
st.subheader("下载")
dl1, dl2, dl3 = st.columns(3)
with dl1:
	if os.path.exists(summary_path):
		with open(summary_path, "rb") as f:
			st.download_button("班级汇总", f.read(), file_name=f"评分汇总_{selected}.xlsx")
with dl2:
	if os.path.exists(personal_dir):
		zb = io.BytesIO()
		with zipfile.ZipFile(zb, "w", zipfile.ZIP_DEFLATED) as zf:
			for f in os.listdir(personal_dir):
				if f.endswith(".xlsx"):
					zf.write(os.path.join(personal_dir, f), f)
		st.download_button("个人明细(ZIP)", zb.getvalue(), file_name=f"个人明细_{selected}.zip")
with dl3:
	if os.path.exists(check_path):
		with open(check_path, "rb") as f:
			st.download_button("查重报告", f.read(), file_name=f"查重报告_{selected}.xlsx")
