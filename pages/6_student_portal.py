"""
学生成绩查询
"""
import streamlit as st
import os, sys, glob, io, zipfile
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from src.ui_style import inject; inject()

if not st.session_state.get("authenticated", False):
    st.warning("请先登录")
    st.stop()

st.title("学生成绩查询")

# ============================================================
#  数据加载
# ============================================================
def list_classes():
    classes = []
    if not os.path.exists("output"):
        return classes
    for d in os.listdir("output"):
        dpath = os.path.join("output", d)
        summary = os.path.join(dpath, f"评分汇总_{d}.xlsx")
        if os.path.isdir(dpath) and os.path.exists(summary):
            classes.append({"name": d, "dir": dpath, "summary": summary})
    return sorted(classes, key=lambda x: x["name"])

classes = list_classes()
if not classes:
    st.info("还没有阅卷记录")
    st.stop()

@st.cache_data(ttl=30)
def load_class_data(summary_path):
    try:
        wb = load_workbook(summary_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(all_rows) < 5:
            return None
        row2, row3, row4 = all_rows[1], all_rows[2], all_rows[3]
        max_len = max(len(row2), len(row3), len(row4))

        columns_info = []
        for i in range(max_len):
            r2 = str(row2[i]).strip() if i < len(row2) and row2[i] else ""
            r3 = str(row3[i]).strip() if i < len(row3) and row3[i] else ""
            r4 = row4[i] if i < len(row4) and row4[i] else ""
            if i < 3:
                columns_info.append(("fixed", None))
                continue
            if "总分" in (r2 + r3):
                columns_info.append(("总分", None))
                break
            c_max = int(r4) if isinstance(r4, (int, float)) and not isinstance(r4, bool) else 0
            label = f"{r2}_{r3}" if r2 and r3 else (r3 or r2 or f"col_{i}")
            columns_info.append((label.replace("\n", " ").strip(), c_max))

        students = []
        for row in all_rows[4:]:
            vals = list(row) if row else []
            if not vals or all(v is None for v in vals):
                continue
            sid = str(vals[1]).strip() if len(vals) > 1 and vals[1] else ""
            sname = str(vals[2]).strip() if len(vals) > 2 and vals[2] else ""
            first_col = str(vals[0]).strip() if vals and vals[0] else ""
            if not sid or first_col in ("", "平均", "平均分", "None", "nan"):
                continue

            scores_detail = []
            total = 0
            for i in range(3, len(vals)):
                if i - 3 < len(columns_info) - 3:
                    info = columns_info[i]
                    if info[0] == "总分":
                        total = float(vals[i]) if vals[i] is not None else 0
                        break
                    score = float(vals[i]) if vals[i] is not None else 0
                    scores_detail.append({
                        "label": info[0], "score": score, "max": info[1] or 0,
                    })

            students.append({
                "sid": sid, "name": sname, "scores": scores_detail, "total": total,
            })

        return {"students": students, "columns": columns_info[3:]}
    except Exception:
        return None

# ============================================================
#  UI
# ============================================================
tab1, tab2 = st.tabs(["学号查询", "全班浏览"])

with tab1:
    col_class, col_sid = st.columns([1, 2])
    with col_class:
        query_class = st.selectbox("选择班级", [c["name"] for c in classes], key="query_class")
    with col_sid:
        query_sid = st.text_input("输入学号", placeholder="例如：255102030101", key="query_sid")

    if query_class and query_sid:
        class_info = next((c for c in classes if c["name"] == query_class), None)
        if class_info:
            cdata = load_class_data(class_info["summary"])
            if cdata:
                student = next((s for s in cdata["students"] if s["sid"] == query_sid.strip()), None)

                if student:
                    st.success(f"找到学生：{student['name']}（学号：{student['sid']}）")

                    st.markdown("---")
                    st.subheader("成绩总览")

                    all_totals = sorted([s["total"] for s in cdata["students"]], reverse=True)
                    rank = all_totals.index(student["total"]) + 1 if student["total"] in all_totals else len(all_totals)
                    avg_total = sum(s["total"] for s in cdata["students"]) / len(cdata["students"])
                    std_dev = (sum((s["total"] - avg_total) ** 2 for s in cdata["students"]) / len(cdata["students"])) ** 0.5
                    z_score = (student["total"] - avg_total) / std_dev if std_dev > 0 else 0
                    max_total = max(s["total"] for s in cdata["students"])

                    metric_cols = st.columns(5)
                    with metric_cols[0]:
                        st.metric("总分", f"{student['total']:.0f}")
                    with metric_cols[1]:
                        st.metric("标准分(Z)", f"{z_score:+.2f}")
                    with metric_cols[2]:
                        st.metric("班级平均", f"{avg_total:.1f}")
                    with metric_cols[3]:
                        st.metric("班级排名", f"{rank} / {len(cdata['students'])}")
                    with metric_cols[4]:
                        st.metric("班级最高", f"{max_total:.0f}")

                    st.progress(
                        1 - (rank - 1) / len(cdata["students"]),
                        text=f"超过了 {len(cdata['students']) - rank} 位同学（共 {len(cdata['students'])} 人）"
                    )

                    st.markdown("---")
                    st.subheader("各题得分明细")
                    if student["scores"]:
                        detail_data = []
                        for sc in student["scores"]:
                            rate = (sc["score"] / sc["max"] * 100) if sc["max"] > 0 else 0
                            detail_data.append({
                                "评分项": sc["label"],
                                "得分": f"{sc['score']:.0f}",
                                "满分": f"{sc['max']:.0f}",
                                "得分率": f"{rate:.0f}%",
                            })
                        st.dataframe(pd.DataFrame(detail_data), use_container_width=True, hide_index=True)

                        for sc in student["scores"]:
                            if sc["max"] > 0:
                                rate = sc["score"] / sc["max"]
                                st.caption(sc["label"][:40])
                                st.progress(float(rate), text=f"{sc['score']:.0f}/{sc['max']:.0f}")

                    st.markdown("---")
                    personal_dir = os.path.join(class_info["dir"], "个人成绩")
                    personal_file = None
                    if os.path.exists(personal_dir):
                        for f in os.listdir(personal_dir):
                            if query_sid.strip() in f and f.endswith(".xlsx"):
                                personal_file = os.path.join(personal_dir, f)
                                break

                    if personal_file and os.path.exists(personal_file):
                        with open(personal_file, "rb") as f:
                            st.download_button(
                                f"下载 {student['name']} 成绩单",
                                f.read(),
                                file_name=f"成绩单_{query_sid}_{student['name']}.xlsx",
                                use_container_width=True,
                            )
                    else:
                        st.caption("暂无个人成绩明细文件")
                else:
                    st.warning(f"未找到学号为「{query_sid}」的学生")
                    sids = [s["sid"] for s in cdata["students"]]
                    st.caption(f"该班级共 {len(sids)} 名学生")

with tab2:
    st.subheader("全班成绩一览")
    browse_class = st.selectbox("选择班级", [c["name"] for c in classes], key="browse_class")
    show_rank = st.checkbox("显示排名", value=True, key="show_rank")

    if browse_class:
        class_info = next((c for c in classes if c["name"] == browse_class), None)
        if class_info:
            cdata = load_class_data(class_info["summary"])
            if cdata and cdata["students"]:
                sorted_students = sorted(cdata["students"], key=lambda s: s["total"], reverse=True)

                table_data = []
                for i, s in enumerate(sorted_students):
                    row_data = {
                        "排名": i + 1 if show_rank else "",
                        "学号": s["sid"],
                        "姓名": s["name"],
                        "总分": f"{s['total']:.0f}",
                    }
                    for sc in s["scores"]:
                        short_label = sc["label"].split("_")[-1] if "_" in sc["label"] else sc["label"]
                        row_data[short_label[:15]] = f"{sc['score']:.0f}/{sc['max']:.0f}"
                    table_data.append(row_data)

                display_df = pd.DataFrame(table_data)
                if not show_rank:
                    display_df = display_df.drop(columns=["排名"])
                st.dataframe(display_df, use_container_width=True, hide_index=True,
                             height=min(600, 35 * len(sorted_students) + 38))

                st.divider()
                st.caption(
                    f"共 {len(sorted_students)} 名学生 | "
                    f"平均分 {sum(s['total'] for s in sorted_students)/len(sorted_students):.1f} | "
                    f"最高 {max(s['total'] for s in sorted_students):.0f} | "
                    f"最低 {min(s['total'] for s in sorted_students):.0f}"
                )
            else:
                st.warning("无法加载成绩数据")
