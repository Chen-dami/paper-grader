"""
报告生成器 —— 个人明细 + 按班级汇总 + 统计。
题目数量从 rubric 动态读取，不硬编码。
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# 样式常量
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
RED_FONT = Font(name="微软雅黑", size=10, color="FF0000")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="center")


def individual_report(student: dict, all_scores: dict, rubric: dict,
                      output_dir: str = "output") -> str:
    """单人生成一张评分明细 Excel"""
    sid = student.get("学号", "unknown")
    sname = student.get("姓名", "unknown")
    sclass = student.get("班级", "")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{sid}_{sname}"

    questions = rubric["questions"]
    max_col = 7

    # 标题区
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws['A1'] = f"《{rubric['exam']['name']}》评分明细"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws['A2'] = f"班级：{sclass}    学号：{sid}    姓名：{sname}"
    ws['A2'].font = NORMAL_FONT
    ws['A2'].alignment = CENTER

    # 表头
    headers = ["题号", "题目", "得分项", "得分", "满分", "得分率", "说明"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    row = 5
    total_score = 0
    total_max = 0

    for q in questions:
        qid = q["id"]
        q_scores = all_scores.get(qid, {})

        for i, c in enumerate(q["criteria"]):
            key = f"得分_{c['id']}_{c['name']}"
            score = q_scores.get(key, 0)
            max_s = c["max"]

            ws.cell(row=row, column=1, value=str(qid) if i == 0 else "").font = NORMAL_FONT
            ws.cell(row=row, column=2, value=q["name"] if i == 0 else "").font = NORMAL_FONT
            ws.cell(row=row, column=3, value=c["name"]).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=score).font = NORMAL_FONT
            ws.cell(row=row, column=5, value=max_s).font = NORMAL_FONT
            rate = score / max_s if max_s > 0 else 0
            ws.cell(row=row, column=6, value=f"{rate:.0%}").font = NORMAL_FONT
            ws.cell(row=row, column=7, value=c["desc"]).font = Font(name="微软雅黑", size=9, color="666666")

            # 着色
            sc = ws.cell(row=row, column=4)
            if max_s > 0:
                if rate >= 0.8:
                    sc.fill = PASS_FILL
                elif rate < 0.5:
                    sc.fill = FAIL_FILL

            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1
            total_score += score
            total_max += max_s

        # 小计
        q_total = sum(q_scores.get(f"得分_{c['id']}_{c['name']}", 0) for c in q["criteria"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=f"  题{qid} 小计").font = BOLD_FONT
        ws.cell(row=row, column=4, value=q_total).font = BOLD_FONT
        ws.cell(row=row, column=5, value=q["max_score"]).font = BOLD_FONT
        rate = q_total / q["max_score"] if q["max_score"] > 0 else 0
        ws.cell(row=row, column=6, value=f"{rate:.0%}").font = BOLD_FONT
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # 总计
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="总  分").font = Font(name="微软雅黑", size=14, bold=True)
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=4, value=total_score).font = Font(name="微软雅黑", size=14, bold=True, color="C00000")
    ws.cell(row=row, column=5, value=total_max).font = Font(name="微软雅黑", size=14, bold=True)
    grade_label = _grade(total_score, total_max)
    ws.cell(row=row, column=6, value=grade_label).font = Font(name="微软雅黑", size=14, bold=True, color="C00000")
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER

    # 评语
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    comments = []
    for q in questions:
        qid = q["id"]
        comment = all_scores.get(qid, {}).get("评语", "")
        if comment:
            comments.append(f"题{qid}：{comment}")
    ws.cell(row=row, column=1, value="评语：" + "；".join(comments)).font = NORMAL_FONT
    ws.cell(row=row, column=1).alignment = WRAP
    ws.row_dimensions[row].height = 40

    # 列宽
    widths = [6, 12, 22, 8, 8, 8, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    personal_dir = os.path.join(output_dir, "个人成绩")
    os.makedirs(personal_dir, exist_ok=True)
    filepath = os.path.join(personal_dir, f"评分明细_{sid}_{sname}.xlsx")
    try:
        wb.save(filepath)
    except PermissionError:
        pass
    return filepath


def class_summary_report(results: list, rubric: dict, output_dir: str,
                         class_name: str) -> str:
    """单个班级的汇总表 —— 得分项列式展示"""
    if not results:
        return ""

    questions = rubric["questions"]
    wb = Workbook()

    # ===== Sheet 1: 得分表 =====
    ws = wb.active
    ws.title = "得分表"

    # 构建列映射
    col_map = []  # [(qid, qname, qmax, [(cid, cname, cmax), ...]), ...]
    for q in questions:
        items = [(c["id"], c["name"], c["max"]) for c in q["criteria"]]
        col_map.append((q["id"], q["name"], q["max_score"], items))

    fixed_cols = 3  # 序号, 学号, 姓名
    criterion_cols = sum(len(items) for _, _, _, items in col_map)
    total_cols = fixed_cols + criterion_cols + 1  # +总分

    total_max = rubric["exam"]["total_score"]

    # ---- 行1: 标题 ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws['A1'] = f"《{rubric['exam']['name']}》{class_name} 得分表"
    ws['A1'].font = TITLE_FONT; ws['A1'].alignment = CENTER

    # ---- 行2-4: 表头 ----
    # 行2: 大题名（合并）
    # 行3: 得分项名
    # 行4: 满分值
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    ws.cell(row=2, column=1, value="序号").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL; ws.cell(row=2, column=1).border = THIN_BORDER; ws.cell(row=2, column=1).alignment = CENTER
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
    ws.cell(row=2, column=2, value="学号").font = HEADER_FONT
    ws.cell(row=2, column=2).fill = HEADER_FILL; ws.cell(row=2, column=2).border = THIN_BORDER; ws.cell(row=2, column=2).alignment = CENTER
    ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
    ws.cell(row=2, column=3, value="姓名").font = HEADER_FONT
    ws.cell(row=2, column=3).fill = HEADER_FILL; ws.cell(row=2, column=3).border = THIN_BORDER; ws.cell(row=2, column=3).alignment = CENTER

    cur_col = fixed_cols + 1
    for qid, qname, qmax, items in col_map:
        n = len(items)
        if n > 1:
            ws.merge_cells(start_row=2, start_column=cur_col, end_row=2, end_column=cur_col + n - 1)
        c = ws.cell(row=2, column=cur_col, value=f"Q{qid} {qname}({qmax}分)")
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER; c.alignment = CENTER
        for j, (cid, cname, cmax) in enumerate(items):
            ws.cell(row=3, column=cur_col + j, value=cname).font = HEADER_FONT
            ws.cell(row=3, column=cur_col + j).fill = HEADER_FILL; ws.cell(row=3, column=cur_col + j).border = THIN_BORDER; ws.cell(row=3, column=cur_col + j).alignment = CENTER
            ws.cell(row=4, column=cur_col + j, value=cmax).font = NORMAL_FONT
            ws.cell(row=4, column=cur_col + j).border = THIN_BORDER; ws.cell(row=4, column=cur_col + j).alignment = CENTER
        cur_col += n

    # 总分
    ws.merge_cells(start_row=2, start_column=cur_col, end_row=4, end_column=cur_col)
    ws.cell(row=2, column=cur_col, value="总分").font = HEADER_FONT
    ws.cell(row=2, column=cur_col).fill = HEADER_FILL; ws.cell(row=2, column=cur_col).border = THIN_BORDER; ws.cell(row=2, column=cur_col).alignment = CENTER
    total_score_col = cur_col

    # ---- 学生数据 ----
    row = 5
    all_criterion_sums = {cid: 0.0 for _, _, _, items in col_map for cid, _, _ in items}
    all_criterion_counts = {cid: 0 for _, _, _, items in col_map for cid, _, _ in items}
    total_sum = 0.0

    for i, r in enumerate(results):
        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT; ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=r.get("student_id", "")).font = NORMAL_FONT; ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=r.get("student_name", "")).font = NORMAL_FONT; ws.cell(row=row, column=3).border = THIN_BORDER

        cur_col = fixed_cols + 1
        student_total = 0
        criteria = r.get("_criteria", {})
        for qid, qname, qmax, items in col_map:
            q_criteria = criteria.get(str(qid), {})
            for cid, cname, cmax in items:
                cd = q_criteria.get(cid, {})
                sc = cd.get("score", 0)
                ws.cell(row=row, column=cur_col, value=sc).font = NORMAL_FONT
                ws.cell(row=row, column=cur_col).alignment = CENTER
                ws.cell(row=row, column=cur_col).border = THIN_BORDER
                if cmax > 0 and sc >= cmax * 0.9:
                    ws.cell(row=row, column=cur_col).fill = PASS_FILL
                elif cmax > 0 and sc < cmax * 0.5:
                    ws.cell(row=row, column=cur_col).fill = FAIL_FILL
                student_total += sc
                all_criterion_sums[cid] = all_criterion_sums.get(cid, 0) + sc
                all_criterion_counts[cid] = all_criterion_counts.get(cid, 0) + 1
                cur_col += 1

        total = r.get("total_score", 0)
        ws.cell(row=row, column=total_score_col, value=total).font = BOLD_FONT
        ws.cell(row=row, column=total_score_col).alignment = CENTER
        ws.cell(row=row, column=total_score_col).border = THIN_BORDER

        if total < 60:
            for c in range(1, total_cols + 1):
                ws.cell(row=row, column=c).font = RED_FONT

        total_sum += total
        row += 1

    # ---- 平均行 ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="  平均").font = BOLD_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER; ws.cell(row=row, column=1).alignment = CENTER
    for c in range(2, 4):
        ws.cell(row=row, column=c).border = THIN_BORDER

    cur_col = fixed_cols + 1
    for qid, qname, qmax, items in col_map:
        for cid, cname, cmax in items:
            cnt = all_criterion_counts.get(cid, 1)
            avg = round(all_criterion_sums.get(cid, 0) / cnt, 1) if cnt > 0 else 0
            ws.cell(row=row, column=cur_col, value=avg).font = BOLD_FONT
            ws.cell(row=row, column=cur_col).alignment = CENTER
            ws.cell(row=row, column=cur_col).border = THIN_BORDER
            cur_col += 1

    total_avg = round(total_sum / len(results), 1) if results else 0
    ws.cell(row=row, column=total_score_col, value=total_avg).font = BOLD_FONT
    ws.cell(row=row, column=total_score_col).alignment = CENTER
    ws.cell(row=row, column=total_score_col).border = THIN_BORDER

    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    for col in range(4, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # ===== Sheet 2: 统计 =====
    ws2 = wb.create_sheet("统计")
    stat_headers = ["指标"] + [q["name"] for q in questions] + ["总分"]
    for col, h in enumerate(stat_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER; c.alignment = CENTER

    # 满分
    ws2.cell(row=2, column=1, value="满分").font = BOLD_FONT
    for qi, q in enumerate(questions):
        ws2.cell(row=2, column=2 + qi, value=q["max_score"]).font = NORMAL_FONT
        ws2.cell(row=2, column=2 + qi).border = THIN_BORDER; ws2.cell(row=2, column=2 + qi).alignment = CENTER
    ws2.cell(row=2, column=2 + len(questions), value=total_max).font = NORMAL_FONT
    ws2.cell(row=2, column=2 + len(questions)).border = THIN_BORDER
    ws2.cell(row=2, column=2 + len(questions)).alignment = CENTER

    # 平均
    ws2.cell(row=3, column=1, value="平均分").font = BOLD_FONT
    q_avgs = []
    for qi, q in enumerate(questions):
        qid = q["id"]
        scores = [r.get(f"q{qid}_score", 0) for r in results]
        avg = round(sum(scores) / len(scores), 1)
        q_avgs.append(avg)
        ws2.cell(row=3, column=2 + qi, value=avg).font = NORMAL_FONT
        ws2.cell(row=3, column=2 + qi).border = THIN_BORDER
        ws2.cell(row=3, column=2 + qi).alignment = CENTER
    ws2.cell(row=3, column=2 + len(questions), value=round(sum(q_avgs), 1)).font = BOLD_FONT
    ws2.cell(row=3, column=2 + len(questions)).border = THIN_BORDER
    ws2.cell(row=3, column=2 + len(questions)).alignment = CENTER

    # 得分率
    ws2.cell(row=4, column=1, value="得分率").font = BOLD_FONT
    for qi, q in enumerate(questions):
        r = q_avgs[qi] / q["max_score"] * 100 if q["max_score"] > 0 else 0
        ws2.cell(row=4, column=2 + qi, value=f"{r:.0f}%").font = NORMAL_FONT
        ws2.cell(row=4, column=2 + qi).border = THIN_BORDER
        ws2.cell(row=4, column=2 + qi).alignment = CENTER
    overall_rate = sum(q_avgs) / total_max * 100 if total_max > 0 else 0
    ws2.cell(row=4, column=2 + len(questions), value=f"{overall_rate:.0f}%").font = BOLD_FONT
    ws2.cell(row=4, column=2 + len(questions)).border = THIN_BORDER
    ws2.cell(row=4, column=2 + len(questions)).alignment = CENTER

    # 分数分布
    ws2.cell(row=6, column=1, value="分数段").font = BOLD_FONT
    ws2.cell(row=6, column=2, value="人数").font = BOLD_FONT
    bands = [
        ("90-100 优秀", lambda s: s >= 90),
        ("80-89 良好", lambda s: 80 <= s < 90),
        ("70-79 中等", lambda s: 70 <= s < 80),
        ("60-69 及格", lambda s: 60 <= s < 70),
        ("<60 不及格", lambda s: s < 60),
    ]
    for bi, (label, fn) in enumerate(bands):
        cnt = sum(1 for r in results if fn(r.get("total_score", 0)))
        ws2.cell(row=7 + bi, column=1, value=label).font = NORMAL_FONT
        ws2.cell(row=7 + bi, column=2, value=cnt).font = NORMAL_FONT
        ws2.cell(row=7 + bi, column=2).alignment = CENTER
        for c in range(1, 3):
            ws2.cell(row=7 + bi, column=c).border = THIN_BORDER

    # 列宽
    for col in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws2.column_dimensions['A'].width = 18
    for col in range(2, len(questions) + 3):
        ws2.column_dimensions[get_column_letter(col)].width = 12

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"评分汇总_{class_name}.xlsx")
    try:
        wb.save(filepath)
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        filepath = os.path.join(output_dir, f"评分汇总_{class_name}_{ts}.xlsx")
        wb.save(filepath)
    return filepath


def print_stats(stats: dict):
    """终端打印统计"""
    if not stats:
        print("\n  (数据库未启用，无统计)\n")
        return
    bands = stats.get('bands', {})
    print(f"""
============================================================
  总份数：{stats.get('total', 0)}    平均分：{stats.get('avg_score', 0)}
  最高分：{stats.get('max_score', 0)}    最低分：{stats.get('min_score', 0)}
  90+：{bands.get('90+', 0)}人  80-89：{bands.get('80-89', 0)}人
  70-79：{bands.get('70-79', 0)}人  60-69：{bands.get('60-69', 0)}人  <60：{bands.get('<60', 0)}人
  Token：{stats.get('tokens_input', 0):,} + {stats.get('tokens_output', 0):,}
============================================================""")


def _grade(score: float, total: float) -> str:
    if total == 0:
        return "N/A"
    pct = score / total * 100
    if pct >= 90:
        return "A 优秀"
    elif pct >= 80:
        return "B 良好"
    elif pct >= 70:
        return "C 中等"
    elif pct >= 60:
        return "D 及格"
    else:
        return "F 不及格"
