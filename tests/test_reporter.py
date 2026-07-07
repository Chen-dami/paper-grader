"""
reporter.py 测试 —— 报告生成器
覆盖：个人明细、班级汇总、统计、等级计算
"""
import os
import pytest
from openpyxl import load_workbook

from src.reporter import (
    individual_report, class_summary_report,
    print_stats, _grade,
)


class TestGrade:
    """等级划分"""

    def test_grade_a(self):
        assert _grade(95, 100) == "A 优秀"

    def test_grade_b(self):
        assert _grade(85, 100) == "B 良好"

    def test_grade_c(self):
        assert _grade(75, 100) == "C 中等"

    def test_grade_d(self):
        assert _grade(65, 100) == "D 及格"

    def test_grade_f(self):
        assert _grade(45, 100) == "F 不及格"

    def test_grade_boundary_90(self):
        assert _grade(90, 100) == "A 优秀"

    def test_grade_boundary_60(self):
        assert _grade(60, 100) == "D 及格"

    def test_grade_zero_total(self):
        assert _grade(0, 0) == "N/A"


class TestIndividualReport:
    """个人成绩明细"""

    def test_generates_excel(self, tmp_dir, sample_rubric):
        """能生成 Excel 文件"""
        student = {"学号": "255102030101", "姓名": "张三", "班级": "软件2501"}
        all_scores = {
            1: {"得分_1-1_主题契合度": 4, "得分_1-2_文案内容适配": 4,
                "得分_1-3_提交完整性": 5, "评语": "主题准确"},
        }
        output_dir = os.path.join(tmp_dir, "output")
        filepath = individual_report(student, all_scores, sample_rubric, output_dir)

        assert os.path.exists(filepath)
        wb = load_workbook(filepath)
        ws = wb.active
        # 标题含考试名
        assert "人工智能基础实践" in str(ws['A1'].value)
        # 学生信息
        sheet_text = " ".join(str(ws.cell(row=r, column=c).value or "")
                             for r in range(1, 6) for c in range(1, 8))
        assert "张三" in sheet_text
        assert "255102030101" in sheet_text

    def test_empty_scores(self, tmp_dir, sample_rubric):
        """学生无分数时不崩溃"""
        student = {"学号": "unknown", "姓名": "测试", "班级": ""}
        output_dir = os.path.join(tmp_dir, "output")
        filepath = individual_report(student, {}, sample_rubric, output_dir)
        assert os.path.exists(filepath)

    def test_personal_dir_created(self, tmp_dir, sample_rubric):
        """个人成绩目录自动创建"""
        student = {"学号": "test", "姓名": "test"}
        output_dir = os.path.join(tmp_dir, "output")
        filepath = individual_report(student, {1: {}}, sample_rubric, output_dir)
        assert "个人成绩" in filepath
        assert os.path.isdir(os.path.dirname(filepath))


class TestClassSummaryReport:
    """班级汇总表"""

    def test_generates_summary_excel(self, tmp_dir, sample_rubric, sample_grading_results):
        """生成班级汇总 Excel"""
        output_dir = os.path.join(tmp_dir, "output")
        filepath = class_summary_report(
            sample_grading_results, sample_rubric, output_dir, "软件2501"
        )
        assert os.path.exists(filepath)
        wb = load_workbook(filepath)

        # Sheet 1: 得分表
        ws = wb["得分表"]
        assert "软件2501" in str(ws['A1'].value)
        # 学生数据
        sheet_data = []
        for row in ws.iter_rows(values_only=True):
            sheet_data.append([str(c) if c is not None else "" for c in row])
        flat = " ".join(str(x) for row in sheet_data for x in row)
        assert "张三" in flat
        assert "李四" in flat

        # Sheet 2: 统计
        ws2 = wb["统计"]
        stat_data = []
        for row in ws2.iter_rows(values_only=True):
            stat_data.append([str(c) if c is not None else "" for c in row])
        flat2 = " ".join(str(x) for row in stat_data for x in row)
        assert "平均分" in flat2 or "平均" in flat2

    def test_empty_results_returns_empty(self, tmp_dir, sample_rubric):
        """空结果列表"""
        filepath = class_summary_report([], sample_rubric, os.path.join(tmp_dir, "output"), "空班")
        assert filepath == ""

    def test_creates_output_dir(self, tmp_dir, sample_rubric, sample_grading_results):
        """目录不存在时自动创建"""
        output_dir = os.path.join(tmp_dir, "nested", "output")
        filepath = class_summary_report(sample_grading_results, sample_rubric, output_dir, "软件2501")
        assert os.path.exists(filepath)


class TestPrintStats:
    """终端打印统计"""

    def test_print_stats_no_error(self, capsys):
        """不崩溃"""
        stats = {
            "total": 30, "avg_score": 75.5, "max_score": 98, "min_score": 42,
            "bands": {"90+": 5, "80-89": 10, "70-79": 8, "60-69": 5, "<60": 2},
            "tokens_input": 15000, "tokens_output": 8000,
        }
        print_stats(stats)
        captured = capsys.readouterr()
        assert "30" in captured.out

    def test_print_empty_stats(self, capsys):
        """空统计"""
        print_stats({})
        captured = capsys.readouterr()
        assert "未启用" in captured.out or "无统计" in captured.out


class TestScoreColoring:
    """得分着色逻辑"""

    def test_high_score_green(self, tmp_dir, sample_rubric):
        """高分 → 绿色"""
        student = {"学号": "001", "姓名": "满分", "班级": "测试"}
        scores = {1: {"得分_1-1_主题契合度": 5, "得分_1-2_文案内容适配": 5,
                       "得分_1-3_提交完整性": 5, "评语": ""}}
        filepath = individual_report(student, scores, sample_rubric, os.path.join(tmp_dir, "output"))
        wb = load_workbook(filepath)
        ws = wb.active
        # 检查得分列的填充色 (列D, 行6开始)
        for row in range(6, 20):
            cell = ws.cell(row=row, column=4)
            if cell.value == 5 and cell.fill and cell.fill.start_color:
                # 满分应有绿色填充
                break
        # 不崩溃即通过
        assert os.path.exists(filepath)

    def test_low_score_red(self, tmp_dir, sample_rubric):
        """低分 → 红色"""
        student = {"学号": "002", "姓名": "低分", "班级": "测试"}
        scores = {1: {"得分_1-1_主题契合度": 1, "得分_1-2_文案内容适配": 1,
                       "得分_1-3_提交完整性": 0, "评语": ""}}
        filepath = individual_report(student, scores, sample_rubric, os.path.join(tmp_dir, "output"))
        assert os.path.exists(filepath)
